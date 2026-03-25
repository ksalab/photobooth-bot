import asyncio
import logging
import os
import re
from datetime import datetime

import pandas as pd
from aiogram import Bot, Dispatcher, F, types
from aiogram.filters import CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import FSInputFile, InlineKeyboardButton, InlineKeyboardMarkup
from dotenv import load_dotenv
from openpyxl.styles import Font, PatternFill

load_dotenv()
TOKEN = os.getenv("BOT_TOKEN")

# Завантаження та конвертація в список int
admin_ids_str = os.getenv("ADMIN_IDS", "")
ADMIN_IDS = [int(id.strip()) for id in admin_ids_str.split(",") if id.strip()]

MANAGER_URL = os.getenv("MANAGER_URL", "")
DB_PATH = os.getenv("DB_PATH", "")
FILE_PATH = os.getenv("FILE_PATH", "")

# Створюємо папку, якщо її немає
if not os.path.exists("db"):
    os.makedirs("db")
if not os.path.exists("assets"):
    os.makedirs("assets")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.FileHandler("bot_log.log"), logging.StreamHandler()],
)
logger = logging.getLogger(__name__)

bot = Bot(token=TOKEN)
dp = Dispatcher()

logger.info(f"Admins: {ADMIN_IDS}")


class Survey(StatesGroup):
    city = State()
    venue_name = State()
    venue_format = State()
    venue_format_custom = State()  # Для варіанту "Інше"
    guests_count = State()
    location_type = State()
    interest_reason = State()
    user_name = State()
    user_contact = State()


# --- Функція для збереження даних у Excel ---
def save_to_excel(data: dict):
    # 1. Формуємо словник так, щоб 'Дата' була першою
    ordered_data = {"Дата": datetime.now().strftime("%d.%m.%Y %H:%M:%S")}
    ordered_data.update(data)

    df_new = pd.DataFrame([ordered_data])

    if os.path.exists(DB_PATH):
        df_old = pd.read_excel(DB_PATH)
        df_final = pd.concat([df_old, df_new], ignore_index=True)
    else:
        df_final = df_new

    # 2. Використовуємо ExcelWriter для стилізації
    with pd.ExcelWriter(DB_PATH, engine="openpyxl") as writer:
        df_final.to_excel(writer, index=False, sheet_name="Leads")

        worksheet = writer.sheets["Leads"]

        # Стилі для заголовка

        header_fill = PatternFill(
            start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
        )
        header_font = Font(bold=True)

        # 3. Фіксуємо перший рядок (заголовки)
        worksheet.freeze_panes = "A2"

        # 3. Ширина колонок (від А до N)
        column_widths = [20, 25, 18, 20, 30, 20, 15, 25, 25, 25, 25, 25, 15, 20]

        for i, width in enumerate(column_widths, start=1):
            col_letter = worksheet.cell(row=1, column=i).column_letter
            worksheet.column_dimensions[col_letter].width = width

            # Стилізуємо кожну комірку першого рядка
            cell = worksheet.cell(row=1, column=i)
            cell.fill = header_fill
            cell.font = header_font


# --- Допоміжні функції для клавіатур ---
def get_format_kb():
    buttons = [
        [
            InlineKeyboardButton(text="Ресторан", callback_data="fmt:Ресторан"),
            InlineKeyboardButton(text="Бар", callback_data="fmt:Бар"),
        ],
        [
            InlineKeyboardButton(text="Кафе", callback_data="fmt:Кафе"),
            InlineKeyboardButton(text="Готель", callback_data="fmt:Готель"),
        ],
        [InlineKeyboardButton(text="Інше ✏️", callback_data="fmt:other")],
    ]
    return InlineKeyboardMarkup(inline_keyboard=buttons)


def get_location_kb():
    buttons = [
        [InlineKeyboardButton(text="На вулиці", callback_data="loc:На вулиці")],
        [
            InlineKeyboardButton(
                text="В основному залі", callback_data="loc:В основному залі"
            )
        ],
        [
            InlineKeyboardButton(
                text="В окремій зоні", callback_data="loc:В окремій зоні"
            )
        ],
        [InlineKeyboardButton(text="Ще не знаю", callback_data="loc:Ще не знаю")],
    ]
    return InlineKeyboardMarkup(inline_keyboard=buttons)


def get_interest_kb():
    buttons = [
        [
            InlineKeyboardButton(
                text="Додатковий дохід", callback_data="int:Додатковий дохід"
            )
        ],
        [
            InlineKeyboardButton(
                text="Маркетинг і контент", callback_data="int:Маркетинг і контент"
            )
        ],
        [
            InlineKeyboardButton(
                text="Обидва варіанти", callback_data="int:Обидва варіанти"
            )
        ],
    ]
    return InlineKeyboardMarkup(inline_keyboard=buttons)


# --- Обробники ---
@dp.message(CommandStart())
async def cmd_start(message: types.Message):
    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="Почати розрахунок 🚀", callback_data="start_survey"
                )
            ]
        ]
    )
    logger.info(f"User {message.from_user.id} started the bot")
    await message.answer(
        "Вітаю!\n\nЯ допоможу за 1 хвилину порахувати потенціал доходу від фотокабінки для вашого закладу.",
        reply_markup=kb,
    )


@dp.callback_query(F.data == "start_survey")
async def start_survey(callback: types.CallbackQuery, state: FSMContext):
    await state.set_state(Survey.city)
    logger.info(f"User {callback.from_user.id} clicked 'start_survey'")
    await callback.message.answer("У якому місті знаходиться ваш заклад?")
    await callback.answer()


@dp.message(Survey.city)
async def process_city(message: types.Message, state: FSMContext):
    # Дозволяємо українські, латинські літери, пробіли та дефіс
    if not re.fullmatch(r"[a-zA-Zа-яА-ЯіїєґІЇЄҐ\s\.\-_]+", message.text):
        return await message.answer(
            "Формат закладу може містити лише літери, цифри, дефіс, крапку та підкреслення (_)."
        )

    await state.update_data(city=message.text)
    logger.info(f"User {message.from_user.id} entered city: {message.text}")
    await state.set_state(Survey.venue_name)
    await message.answer("Як називається ваш заклад?")


@dp.message(Survey.venue_name)
async def process_venue_name(message: types.Message, state: FSMContext):
    await state.update_data(venue_name=message.text)
    await state.set_state(Survey.venue_format)
    logger.info(f"User {message.from_user.id} entered venue name: {message.text}")
    await message.answer("Який у вас формат закладу?", reply_markup=get_format_kb())


@dp.callback_query(Survey.venue_format, F.data.startswith("fmt:"))
async def process_format(callback: types.CallbackQuery, state: FSMContext):
    choice = callback.data.split(":")[1]

    if choice == "other":
        await state.set_state(Survey.venue_format_custom)
        await callback.message.answer("Напишіть, будь ласка, ваш формат закладу:")
    else:
        await state.update_data(venue_format=choice)
        await state.set_state(Survey.guests_count)
        await callback.message.answer(
            "Скільки гостей у середньому у вас буває за місяць? (Введіть число)"
        )
    logger.info(f"User {callback.from_user.id} selected format: {choice}")

    await callback.answer()


@dp.message(Survey.venue_format_custom)
async def process_custom_format(message: types.Message, state: FSMContext):
    # Дозволяємо літери, цифри, пробіли та дефіс
    if not re.fullmatch(r"[a-zA-Zа-яА-ЯіїєґІЇЄҐ0-9\s\-]+", message.text):
        return await message.answer(
            "Формат закладу може містити лише літери, цифри та дефіс."
        )

    await state.update_data(venue_format=message.text)
    logger.info(f"User {message.from_user.id} entered custom format: {message.text}")
    await state.set_state(Survey.guests_count)
    await message.answer(
        "Скільки гостей у середньому у вас буває за місяць? (Введіть число)"
    )


@dp.message(Survey.guests_count)
async def process_guests(message: types.Message, state: FSMContext):
    if not message.text.isdigit():
        return await message.answer(
            "Будь ласка, введіть число цифрами (наприклад: 500)."
        )

    await state.update_data(guests_count=int(message.text))
    logger.info(f"User {message.from_user.id} entered guests count: {message.text}")
    await state.set_state(Survey.location_type)
    await message.answer(
        "Де потенційно може стояти фотокабінка?",
        reply_markup=get_location_kb(),
    )


@dp.callback_query(Survey.location_type, F.data.startswith("loc:"))
async def process_location(callback: types.CallbackQuery, state: FSMContext):
    await state.update_data(location_type=callback.data.split(":")[1])
    logger.info(
        f"User {callback.from_user.id} selected location: {callback.data.split(':')[1]}"
    )
    await state.set_state(Survey.interest_reason)
    await callback.message.answer("Що вам цікавіше?", reply_markup=get_interest_kb())
    await callback.answer()


@dp.callback_query(Survey.interest_reason, F.data.startswith("int:"))
async def process_interest(callback: types.CallbackQuery, state: FSMContext):
    await state.update_data(interest_reason=callback.data.split(":")[1])
    logger.info(
        f"User {callback.from_user.id} selected interest: {callback.data.split(':')[1]}"
    )
    await state.set_state(Survey.user_name)
    await callback.message.answer("Як Вас звати?")
    await callback.answer()


@dp.message(Survey.user_name)
async def process_user_name(message: types.Message, state: FSMContext):
    # Тільки літери та пробіли (якщо ім'я подвійне)
    if not re.fullmatch(r"[a-zA-Zа-яА-ЯіїєґІЇЄҐ\s]+", message.text):
        return await message.answer(
            "Будь ласка, введіть ім'я коректно (використовуйте лише літери)."
        )

    await state.update_data(user_name=message.text)
    logger.info(f"User {message.from_user.id} entered name: {message.text}")
    await state.set_state(Survey.user_contact)
    kb = types.ReplyKeyboardMarkup(
        keyboard=[
            [types.KeyboardButton(text="Поділитися номером 📱", request_contact=True)]
        ],
        resize_keyboard=True,
        one_time_keyboard=True,
    )
    await message.answer(
        "Вкажіть ваш телефон або Telegram для зв’язку", reply_markup=kb
    )


@dp.message(Survey.user_contact)
async def process_final(message: types.Message, state: FSMContext):
    contact = message.contact.phone_number if message.contact else message.text
    data = await state.get_data()

    # Розрахунки
    guests = data["guests_count"]
    price = 150
    res_7 = int(guests * 0.07 * price)
    res_10 = int(guests * 0.10 * price)
    res_15 = int(guests * 0.15 * price)

    final_text = (
        f"📋 <b>Ваш прогноз потенціалу для закладу «{data['venue_name']}»</b>\n\n"
        f"{data['user_name']}, при потоці {guests} гостей/місяць фотокабінка може генерувати:\n\n"
        f"🔹 Консервативний сценарій: <b>{res_7} грн/міс</b>\n"
        f"🔸 Реалістичний сценарій: <b>{res_10} грн/міс</b>\n"
        f"🚀 Потенційний сценарій: <b>{res_15} грн/міс</b>\n\n"
        f"Це орієнтовний прогноз при вартості фото 150 грн.\n\n"
        "Фактичний результат залежить від:\n"
        "- Розташування фотокабінки\n"
        "- Атмосфери закладу\n"
        "- Типу аудиторії\n"
        "- Додаткових акцій та інтеграцій\n"
        "- Вартості фото\n\n"
        "<b>Важливо</b>: цей розрахунок показує тільки орієнтовний прямий дохід від фото.\n\n"
        "Але є ще додатковий дохід, який часто дає навіть більше 👇\n\n"
        "📍 <b>1. ДОДАТКОВИЙ ТРАФІК</b>\n"
        "Гості приходять не тільки за їжею чи напоями.\n"
        "<b>Фотокабінка стає причиною прийти саме до вас</b>.\n"
        "👉 ви отримуєте нових гостей без реклами\n"
        "\n\n"
        "📲 <b>2. БЕЗКОШТОВНИЙ МАРКЕТИНГ</b>\n"
        "Кожне фото = контент у соцмережах\n"
        "У середньому:\n"
        "- 1 фото бачать <b>50 – 200 людей</b>\n"
        "- 100 фото = <b>5 000 – 20 000 переглядів</b>\n"
        "👉 це реклама, за яку ви не платите\n"
        "\n\n"
        "🎯 <b>3. ІНСТРУМЕНТ ЛОЯЛЬНОСТІ</b>\n"
        "Фотокабінка легко інтегрується в систему:\n"
        "- QR-знижки\n"
        "- “купи коктейль → отримай фото”\n"
        "- бонуси для гостей\n"
        "👉 стимулює повторні візити\n"
        "\n\n"
        "📸 <b>4. ДРУКОВАНЕ ФОТО = ТРЕНД</b>\n"
        "Люди втомились від цифрового.\n"
        "<b>Фізичне фото = емоція, яку можна забрати з собою</b>.\n"
        "👉 це повернення мейнстріму\n"
        "\n\n"
        "❤️ <b>5. ЕМОЦІЯ = ПОВЕРНЕННЯ</b>\n"
        "Фотокабінка = місце, де:\n"
        "- збираються люди\n"
        "- знайомляться\n"
        "- проводять час\n"
        "👉 це підсилює атмосферу закладу\n"
        "\n\n"
        "<b>Дізнатися більше та отримати персональний бонус $300-500 ( до 1.04.2026 )</b>"
    )

    final_kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="Зв’язатися з менеджером", url=MANAGER_URL)],
            [
                InlineKeyboardButton(
                    text="Отримати каталог 📚", callback_data="get_catalog"
                )
            ],
        ]
    )

    await message.answer(final_text, reply_markup=final_kb, parse_mode="HTML")

    # ПІСЛЯ отримання contact та data, додаємо збереження:
    full_lead_data = {
        "Ім'я": data.get("user_name"),
        "Контакт": contact,
        "Місто": data.get("city"),
        "Назва закладу": data.get("venue_name"),
        "Формат": data.get("venue_format"),
        "К-ть гостей": data.get("guests_count"),
        "Місце встановлення": data.get("location_type"),
        "Ціль": data.get("interest_reason"),
        "Консервативний": f"{res_7} грн",
        "Реалістичний": f"{res_10} грн",
        "Потенційний": f"{res_15} грн",
        "User ID": message.from_user.id,
        "Username": f"@{message.from_user.username}"
        if message.from_user.username
        else "Немає",
    }

    try:
        save_to_excel(full_lead_data)
        logger.info(f"Full data for user {message.from_user.id} saved to {DB_PATH}")
    except Exception as e:
        logger.error(f"Excel saving error: {e}")

    logger.info(f"Data for user {message.from_user.id} saved to Excel")

    # Оновлюємо Inline-клавіатуру для адміна
    admin_kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="Завантажити базу (.xlsx) 📁", callback_data="download_db"
                )
            ]
        ]
    )

    admin_report = (
        f"💰 НОВИЙ РОЗРАХУНОК\n\n"
        f"🏠 Заклад: {data['venue_name']}\n"
        f"📍 Місто: {data['city']}\n"
        f"🏷 Формат: {data['venue_format']}\n"
        f"👥 Гостей на місяць: {guests}\n"
        f"📍 Розташування: {data['location_type']}\n"
        f"🎯 Ціль: {data['interest_reason']}\n\n"
        f"👤 Ім'я: {data['user_name']}\n"
        f"📞 Контакт: {contact}\n\n"
        f"Розраховані сценарії:\n"
        f"🔹 Консервативний сценарій 7% : {res_7} грн/міс\n"
        f"🔸 Реалістичний сценарій 10% : {res_10} грн/міс\n"
        f"🚀 Потенційний сценарій 15% : {res_15} грн/міс\n"
    )

    # Повідомлення адміну
    if ADMIN_IDS:
        for admin_id in ADMIN_IDS:
            try:
                await bot.send_message(
                    admin_id, admin_report, parse_mode="HTML", reply_markup=admin_kb
                )
            except Exception as e:
                logger.error(f"Не вдалося надіслати звіт адміну {admin_id}: {e}")

    logger.info(f"User {message.from_user.id} completed survey. Contact: {contact}")

    await state.clear()


@dp.callback_query(F.data == "download_db")
async def send_db_file(callback: types.CallbackQuery):
    # Логування для відладки: ви побачите в консолі, хто стукає і хто в списку
    logger.info(
        f"Спроба завантаження бази. User ID: {callback.from_user.id}, Admins: {ADMIN_IDS}"
    )

    # Перевірка: чи ID користувача є в списку ADMIN_IDS
    if callback.from_user.id not in ADMIN_IDS:
        logger.warning(f"Access denied for {callback.from_user.id}")
        await callback.answer("Доступ заборонено", show_alert=True)
        return

    if os.path.exists(DB_PATH):
        document = FSInputFile(DB_PATH)
        await callback.message.answer_document(
            document=document,
            caption=f"Актуальна база лідів на {datetime.now().strftime('%d.%m %H:%M')}",
        )
    else:
        await callback.answer("Файл бази ще не створено", show_alert=True)
    await callback.answer()


@dp.callback_query(F.data == "get_catalog")
async def send_catalog(callback: types.CallbackQuery):
    try:
        # Створюємо об'єкт файлу
        document = FSInputFile(FILE_PATH)

        # Надсилаємо документ
        await callback.message.answer_document(
            document=document,
            caption="📚 Ось ваш каталог фотокабінок з детальним описом та цінами.\n\n"
            "Для персональної консультації та отримання бонусу тисніть кнопку нижче 👇",
            reply_markup=InlineKeyboardMarkup(
                inline_keyboard=[
                    [
                        InlineKeyboardButton(
                            text="Зв’язатися з менеджером",
                            url=MANAGER_URL,
                        )
                    ]
                ]
            ),
        )
        logger.info(f"Catalog sent to user {callback.from_user.id}")

    except Exception as e:
        logger.error(f"Error sending catalog: {e}")
        await callback.message.answer(
            "Вибачте, сталася помилка при завантаженні каталогу. Будь ласка, зверніться до менеджера."
        )

    await callback.answer()


async def main():
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
